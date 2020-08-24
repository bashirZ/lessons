import requests
from lxml import etree
import lxml.html
from openpyxl import Workbook
import time

HEADERS = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.105 Safari/537.36','accept':'*/*'}
HOST = 'https://stroymarket-05.ru'
  

def parse(url):
    try:
        api = requests.get(url, headers=HEADERS)
    except:
        return
    tree = lxml.html.document_fromstring(api.text)
    nazv = tree.xpath('//div[@class="item-title"]/a/span/text()')
    foto = tree.xpath('//img[@class="img-responsive "]/@src')
    return (nazv, foto)



    


def main():
    url = "https://stroymarket-05.ru/catalog/03_instrumenty_i_oborudovanie/kliningovoe_oborudovanie/moyki_vysokogo_davleniya/?PAGEN_1={param}"
    
    s = 1
    wb = Workbook()
    ws = wb.active
    while s <=2:
        nazv, foto = parse(url.format(param=s))
        for i in nazv:
            cell = ws.cell(row = nazv.index(i)+1+(s-1)*20, column = 1)
            cell.value = i
        for j in foto:    
            cell = ws.cell(row = foto.index(j)+1+(s-1)*20, column = 2)
            cell.value = HOST+j


        time.sleep(1)
        s += 1
    wb.save('sm05.xls')



if __name__ == "__main__":
    main()