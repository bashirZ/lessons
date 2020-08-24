from pycoder.items import PycoderItem
import scrapy
from openpyxl import load_workbook
from urllib.parse import urljoin
from itertools import count
HOST = 'https://stroymarket-05.ru'
c = 0
def func():
    global c
    c += 1
    return c

class PycoderSpider(scrapy.Spider):
    name = "pycoder"
    start_urls = [
        'https://stroymarket-05.ru/catalog/02_elektrotovary/pribory_ucheta_i_izmereniya/',
    ]
    visited_urls = []
    counter1 = count(1)

    def parse(self, response):
        if response.url not in self.visited_urls:
            self.visited_urls.append(response.url)
            for post_link in response.xpath('//div[@class="item-title"]/a/@href').extract():
                url = urljoin(response.url, post_link)
                yield response.follow(url, callback=self.parse_post)

            next_pages = response.xpath('//li[@class="flex-nav-next "]/a/@href').extract()
            next_page = next_pages[-1]

            next_page_url = urljoin(response.url+'/', next_page)
            yield response.follow(next_page_url, callback=self.parse)

    def parse_post(self, response):
        i = func()
        item = PycoderItem()
        title = response.xpath('//h1/text()').extract()
        item['title'] = title
        foto = response.xpath('//li[@id="photo-0"]/link/@href').extract()
        item['foto'] = foto   
        yield item
        wb = load_workbook("sm05.xlsx")
        ws = wb.worksheets[0]
        nazv = item ['title']
        foto_url = item ['foto']
        for n in nazv:
            cell = ws.cell(row = i, column = 1)
            cell.value = n
        for f in foto_url:    
            cell = ws.cell(row = i, column = 2)
            cell.value = HOST+f
        wb.save('sm05.xlsx')
        




